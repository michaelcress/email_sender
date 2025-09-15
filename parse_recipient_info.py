#!/usr/bin/env python3

"""
parse_recipient_info.py

This script parses the input and emits values appropriate for consumption by the sender script.

"""


import argparse
import csv
import json
import os
import sys
import logging
from typing import List, Dict, Any, Optional, Union
from logging.handlers import TimedRotatingFileHandler
from datetime import datetime, timezone, timedelta

from typing import TypedDict, Literal, Union

import json, time, random, subprocess
from concurrent.futures import ThreadPoolExecutor, as_completed

from m365_oauth_tokeninfo import OAuthToken

from pathlib import Path
from jinja2 import Environment, FileSystemLoader, StrictUndefined, select_autoescape

import tempfile

# logging.basicConfig(
#     level=logging.INFO,
#     format="%(asctime)s %(threadName)s %(levelname)s %(message)s",
#     stream=sys.stdout,
# )

# log = logging.getLogger("email-sender")


def setup_logging(
    log_dir: str | Path = "logs",
    app_name: str = "email-sender",
    level: int = logging.INFO,
    rotate_daily: bool = True,
    keep: int = 14,   # how many days of logs to keep when rotating
):
    """
    Configure root logger to write to console and a date-stamped file.
    If rotate_daily=True, you'll get a new file each midnight with a YYYY-MM-DD suffix.
    If rotate_daily=False, you'll get a single file stamped with today's date when the program starts.
    """
    log_dir = Path(log_dir)
    log_dir.mkdir(parents=True, exist_ok=True)

    # Base formatter used by both console and file
    fmt = "%(asctime)s | %(levelname)s | %(name)s | %(message)s"
    datefmt = "%Y-%m-%d %H:%M:%S"
    formatter = logging.Formatter(fmt, datefmt=datefmt)

    root = logging.getLogger()
    root.setLevel(level)
    # Prevent duplicate handlers on repeated setup calls
    root.handlers.clear()

    # --- Console handler ---
    ch = logging.StreamHandler(stream=sys.stdout)
    ch.setLevel(level)
    ch.setFormatter(formatter)
    root.addHandler(ch)

    # --- File handler ---
    if rotate_daily:
        # Rotates at midnight; filenames like logs/app.log.2025-09-15
        fh = TimedRotatingFileHandler(
            filename=log_dir / f"{app_name}.log",
            when="midnight",
            interval=1,
            backupCount=keep,
            encoding="utf-8",
            utc=False,  # set True if you prefer UTC-based rollover
        )
        fh.suffix = "%Y-%m-%d"  # date-stamped suffix
    else:
        # Single file per run, stamped with today's date
        today = datetime.now().strftime("%Y-%m-%d")
        fh = logging.FileHandler(log_dir / f"{app_name}-{today}.log", encoding="utf-8")

    fh.setLevel(level)
    fh.setFormatter(formatter)
    root.addHandler(fh)

    return root  # optional, in case you want to inspect or add more handlers


log = setup_logging(log_dir="logs", app_name="email-sender", level=logging.DEBUG, rotate_daily=True)


TEMPLATE_DIR = Path(".")
env = Environment(
    loader=FileSystemLoader(TEMPLATE_DIR),
    autoescape=select_autoescape(["html", "xml"]),
    undefined=StrictUndefined,  # fail fast on missing keys
    trim_blocks=True,
    lstrip_blocks=True,
)


def render_email(template_base: str, context: dict) -> str:
    body_tpl = env.get_template(f"{template_base}")
    html = body_tpl.render(**context)
    return html


# -------- CSV --------
def read_csv_to_rows(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [dict(row) for row in reader]

# -------- Excel (.xlsx / .xlsm) --------
def read_excel_to_rows(path: str, sheet: Optional[str] = None) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "Reading Excel files requires the 'openpyxl' package. Install with: pip install openpyxl"
        ) from e

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return []

    headers = [str(h) if h is not None else "" for h in headers]
    out: List[Dict[str, Any]] = []
    for row in rows_iter:
        record = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        out.append(record)
    return out

def to_keyed_dict(rows: List[Dict[str, Any]], key_column: str) -> Dict[str, Dict[str, Any]]:
    mapping: Dict[str, Dict[str, Any]] = {}
    for i, r in enumerate(rows, start=1):
        key = r.get(key_column)
        if key is None:
            # Normalize None to empty string to avoid TypeError on dict keys
            key = ""
        key = str(key)
        mapping[key] = r  # last one wins on duplicate keys
    return mapping

def infer_file_type(path: str) -> Optional[str]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return "csv"
    if ext in {".xlsx", ".xlsm"}:
        return "excel"
    if ext == ".xls":
        return "xls"  # not supported here
    return None


def token_expired(token: Dict, skew_seconds: int = 600) -> bool:
    """
    Returns True if the token is expired or will expire within `skew_seconds`.
    Expects JSON fields: obtained_at (unix seconds) and expires_in (seconds).
    """
    
    return token.seconds_until_expiry < skew_seconds



def format_address_html(address: str) -> str:
    return (
        address
        .replace("_x000D_", "")      # remove artifacts
        .replace("\r\n", "\n")       # normalize CRLF
        .replace("\r", "\n")
        .replace("\n", "<br>")
    )





MAX_CONCURRENCY = 12            # threads in parallel (tune carefully)
RATE_PER_MINUTE = 600           # hard cap across all threads
SPACING = 60.0 / max(1, RATE_PER_MINUTE)


class SendSuccess(TypedDict):
    email: str
    firstname: str
    lastname: str
    country: str
    ok: Literal[True]
    ms: int

class SendError(TypedDict):
    email: str
    firstname: str
    lastname: str
    country: str
    ok: Literal[False]
    error: str

SendResult = Union[SendSuccess, SendError]



def send_one( test_mode:bool, tokenstr: str, acct_username: str, fromname:str, fromaddr: str, subject: str, templateemailhtmlpath: str, rec: Dict ) -> SendResult:
    # light jitter to avoid thundering herd
    time.sleep(random.uniform(0, SPACING))

    address = rec.get("address", "")
    if address is not None:
        address = format_address_html( address )

    # Build context from your row fields
    ctx = {
        "title": rec.get("title", ""),
        "firstname": rec.get("firstname", ""),
        "lastname":  rec.get("lastname", ""),
        "country":  rec.get("entitynamelong", ""),
        "address":  address,
    }

    html = render_email(templateemailhtmlpath, ctx)

    # Write body to a temp file (unique per thread/recipient)
    with tempfile.NamedTemporaryFile("w", delete=False, suffix=".html") as tmp:
        tmp.write(html)
        html_path = tmp.name


    if test_mode is False:
        toAddr = rec.get("EmailsToUse", "")
    else:
        toAddr = "mikecress+wafuniftest@gmail.com"
        # toAddr = "bucurlili13+wafuniftest@gmail.com"
        # toAddr = "wafunif@wafunif.org"

        
    log.info( f"Sending e-mail to: {rec['title']} {rec['firstname']} {rec['lastname']} to {toAddr}" )

    cmd = [
        "build/email-sender",
        "--from_name", fromname,
        "--from", fromaddr,
        "--to", toAddr,
        "--subject", subject,
        "--username", acct_username,
        "--file", html_path,
        "--token", tokenstr
    ]

    backoff = 1.0
    for attempt in range(1, 6):  # up to 5 tries
        t0 = time.time()
        p = subprocess.run(cmd, capture_output=True, text=True)
        dt = time.time() - t0

        if p.returncode == 0:
            return {"email": toAddr, "firstname": rec['firstname'], "lastname": rec['lastname'], "country": rec["entitynamelong"], "ok": True, "ms": int(dt*1000)}
        # transient? back off and retry
        time.sleep(backoff + random.uniform(0, 0.200))
        backoff = min(backoff * 2, 16.0)

    return {
        "email": toAddr,
        "firstname": rec['firstname'],
        "lastname": rec['lastname'],
        "country": rec["entitynamelong"],
        "ok": False,
        "error": (p.stderr or p.stdout).strip()[:2000]
    }


def run_mail_merge( test_mode: bool, token: str, acct_username: str, fromname: str, fromaddr: str, subject: str, templateemailhtmlpath: str, records: Dict):
    results = []
    with ThreadPoolExecutor(max_workers=MAX_CONCURRENCY) as pool:

        #debug
        if test_mode:
            futures = [pool.submit(send_one, test_mode, token.access_token, acct_username, fromname, fromaddr, subject, templateemailhtmlpath, r) for r in [ records[0], records[1], records[2], records[3], records[4], records[5], records[6] ] ]
        
        #production
        else:
            futures = [pool.submit(send_one, test_mode, token.access_token, acct_username, fromname, fromaddr, subject, templateemailhtmlpath, r) for r in records]
        
            
        for fut in as_completed(futures):
            res = fut.result()
            log.info(json.dumps(res, ensure_ascii=False))
            results.append(res)
    return results

def main(argv: Optional[List[str]] = None) -> int:


    parser = argparse.ArgumentParser(description="Read a CSV or Excel spreadsheet into dict structures.")
    parser.add_argument("path", help="Path to the CSV or Excel file.")
    parser.add_argument("--tokenfile", help="File containing the OAuth token for authentication to the Microsoft 365 SMTP server. (Obtained by running m365_token_helper.py)")
    parser.add_argument("--subject", help="Subject of the e-mail to be sent")
    parser.add_argument("--username", help="Username of the Microsoft 365 account that will be sending the e-mail")
    parser.add_argument("--from_name", help="The name of the sender that will be displayed in the From: field")
    parser.add_argument("--from_addr", help="The From address")
    parser.add_argument("--email_template", help="The E-mail template")

    group = parser.add_mutually_exclusive_group()
    group.add_argument(
        "--test",
        dest="test",
        action="store_true",
        help="Only sends email to mikecress+wafuniftest@gmail.com",
    )
    group.add_argument(
        "--no-test",
        dest="test",
        action="store_false",
        help="Send email normally",
    )

    parser.set_defaults(test=True)

    group = parser.add_mutually_exclusive_group()
    group.add_argument("--type", choices=["csv", "excel"], help="Explicitly set the file type.")
    parser.add_argument("--sheet", help="Excel sheet name (defaults to the first/active sheet).", default=None)
    parser.add_argument("--output", "-o", help="If set, write program output to this path.", default=None)
    args = parser.parse_args(argv)

    if args.tokenfile is None:
        log.info("Please supply a valid token file value.", file=sys.stderr)
        return 2

    if args.subject is None:
        log.info("Please supply a valid e-mail subject value.", file=sys.stderr)
        return 2

    if args.username is None:
        log.info("Please supply a valid username value.", file=sys.stderr)
        return 2

    if args.from_name is None:
        log.info("Please supply a valid from_name value.", file=sys.stderr)
        return 2

    if args.from_addr is None:
        log.info("Please supply a valid from_addr value.", file=sys.stderr)
        return 2

    if args.email_template is None:
        log.info("Please supply a valid email template value.", file=sys.stderr)
        return 2

    if args.type is None:
        log.info("Please pass --type {csv,excel} or --excel.", file=sys.stderr)
        return 2

    if args.type == "excel" and infer_file_type(args.path) == "xls":
        log.info(".xls is not supported by this script. Save as .xlsx or .csv, or install a library that supports .xls.", file=sys.stderr)
        return 2

    test_mode = args.test

    try:
        with open(f"{args.tokenfile}", "r", encoding="utf-8") as f:
            token_data = json.load(f)
    except FileNotFoundError:
        log.info(f"{args.tokenfile} not found")
        return 2
    except IsADirectoryError:
        log.info(f"{args.tokenfile} is a directory, not a file")
        return 2
    except PermissionError:
        log.info(f"no permission to read {args.tokenfile}")
        return 2

    
    tok = OAuthToken(**token_data)

    # log.info(f"Token data is {tok}\n")

    has_token_expired = token_expired(tok)

    if has_token_expired:
        log.info("OAuth bearer token has expired. Please refresh it and then retry this script. See README.txt for instructions on how to do this.\n")
        return 2
    

    try:
        if args.type == "csv":
            rows = read_csv_to_rows(args.path)
        elif args.type == "excel":
            rows = read_excel_to_rows(args.path, sheet=args.sheet)
        else:
            log.info(f"Unsupported file type: {args.type}", file=sys.stderr)
            return 2
    except Exception as e:
        log.info(f"Error reading file: {e}", file=sys.stderr)
        return 1

    # for row in rows:
        # log.info( "", row['firstname'], row['lastname'], row['EmailsToUse'] )

    # subject = "WAFUNIF Test"
    # fromname = "WAFUNIF Test"
    # fromaddr = "membership@wafunif.org"
    # templateemailhtmlpath = "testemail2.html.j2"
    
    run_mail_merge( test_mode, tok, args.username, args.from_name, args.from_addr, args.subject, args.email_template, rows )


    return 0

if __name__ == "__main__":
    raise SystemExit(main())
